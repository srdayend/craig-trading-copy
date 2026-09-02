from __future__ import annotations

import csv
import datetime as dt
import json
import re
import urllib.request
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
MANUAL_XLSX = ROOT / "크레이그 매매 수동 저장파일.xlsx"
DETAILS_JSON = ROOT / "data" / "source" / "craig_youtube" / "details.json"
OUT_DIR = ROOT / "data" / "processed" / "gold_context_trades"
PLAYLIST_URL = "https://www.youtube.com/playlist?list=PLEJ3Z6QWT7h3SFJidwxSzzFf-HLJ5biq-"


SCHEMA_COLUMNS = [
    "trade_id",
    "status",
    "source_kind",
    "video_id",
    "video_url",
    "playlist_index_oldest_first",
    "youtube_anchor",
    "market_time_note",
    "market_date",
    "symbol",
    "direction",
    "trade_sequence",
    "pre_entry_thesis_ko",
    "intention_timeline_ko",
    "setup_structure_ko",
    "entry_plan_ko",
    "stop_plan_ko",
    "target_plan_ko",
    "execution_ko",
    "management_ko",
    "result_ko",
    "source_anchors_ko",
    "missing_or_uncertain_ko",
    "rule_features_ko",
    "original_notes_ko",
]


def parse_video_id(url: str) -> str:
    parsed = urlparse(url)
    if parsed.hostname and "youtu.be" in parsed.hostname:
        return parsed.path.strip("/")
    return parse_qs(parsed.query).get("v", [""])[0]


def parse_playlist_index(url: str) -> str:
    return parse_qs(urlparse(url).query).get("index", [""])[0]


def normalize_symbol(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def youtube_anchor_from_excel(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        value = value.time()
    if isinstance(value, dt.time):
        total = value.hour * 60 + value.minute
        return f"{total // 60:02d}:{total % 60:02d}"
    if isinstance(value, dt.timedelta):
        total_minutes = value.days * 24 + value.seconds // 3600
        total_seconds = (value.seconds % 3600) // 60
        total = total_minutes * 60 + total_seconds
        return f"{total // 60:02d}:{total % 60:02d}"
    text = str(value).strip()
    match = re.search(r"(?:(\d+)\s+day,\s*)?(\d{1,2}):(\d{2})(?::(\d{2}))?", text, re.I)
    if not match:
        return text
    days = int(match.group(1) or 0)
    hour_or_min = int(match.group(2))
    minute_or_sec = int(match.group(3))
    sec = int(match.group(4) or 0)
    total = (days * 24 + hour_or_min) * 60 + minute_or_sec
    if match.group(4) is not None:
        total = hour_or_min * 60 + minute_or_sec
        total += days * 24 * 60
        total += 1 if sec >= 30 else 0
    return f"{total // 60:02d}:{total % 60:02d}"


def extract_time_mentions(notes: str) -> str:
    matches = re.findall(r"(?:\d+\s*day,\s*)?\d{1,2}:\d{2}(?::\d{2})?", notes or "", flags=re.I)
    return "; ".join(dict.fromkeys(matches))


def rough_direction(notes: str) -> str:
    text = notes or ""
    long_score = sum(text.count(token) for token in ["롱", "매수", "상승", "long"])
    short_score = sum(text.count(token) for token in ["숏", "매도", "하락", "short"])
    if long_score > short_score:
        return "long"
    if short_score > long_score:
        return "short"
    return "unknown_until_frame_review"


def build_manual_seed() -> None:
    wb = openpyxl.load_workbook(MANUAL_XLSX, data_only=True)
    ws = wb.active
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        video_url = ws.cell(row_idx, 1).value
        if not video_url:
            continue
        market_time_note = ws.cell(row_idx, 2).value or ""
        youtube_raw = ws.cell(row_idx, 3).value or ""
        symbol = normalize_symbol(ws.cell(row_idx, 4).value)
        notes = ws.cell(row_idx, 5).value or ""
        video_id = parse_video_id(str(video_url))
        playlist_index = parse_playlist_index(str(video_url))
        anchor = youtube_anchor_from_excel(youtube_raw)
        anchor_slug = anchor.replace(":", "m") + "s" if anchor else "no_anchor"
        trade_id = f"{video_id}_manual_row{row_idx:02d}_{anchor_slug}"
        anchors = "; ".join(x for x in [f"main={anchor}" if anchor else "", extract_time_mentions(notes)] if x)
        rows.append(
            {
                "trade_id": trade_id,
                "status": "needs_frame_review",
                "source_kind": "manual_seed",
                "video_id": video_id,
                "video_url": str(video_url),
                "playlist_index_oldest_first": playlist_index,
                "youtube_anchor": anchor,
                "market_time_note": str(market_time_note),
                "market_date": "",
                "symbol": symbol,
                "direction": rough_direction(notes),
                "trade_sequence": "",
                "pre_entry_thesis_ko": "",
                "intention_timeline_ko": "",
                "setup_structure_ko": "",
                "entry_plan_ko": "",
                "stop_plan_ko": "",
                "target_plan_ko": "",
                "execution_ko": "",
                "management_ko": "",
                "result_ko": "",
                "source_anchors_ko": anchors,
                "missing_or_uncertain_ko": "사용자 수동 메모 기반 seed. 프레임/포지션박스/recap 대조 후 gold 승격 가능 여부 판단 필요.",
                "rule_features_ko": "",
                "original_notes_ko": str(notes),
            }
        )

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / "manual_seed_contexts.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SCHEMA_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def fetch_playlist_indices() -> dict[str, int]:
    req = urllib.request.Request(PLAYLIST_URL, headers={"User-Agent": "Mozilla/5.0"})
    html = urllib.request.urlopen(req, timeout=20).read().decode("utf-8", "replace")
    pairs = re.findall(r"watch\?v=([A-Za-z0-9_-]{11}).{0,220}?index=(\d+)", html)
    return {video_id: int(index) for video_id, index in pairs}


def build_review_scope() -> None:
    details = json.loads(DETAILS_JSON.read_text(encoding="utf-8"))
    by_id = {row["id"]: row for row in details}
    playlist_indices = fetch_playlist_indices()

    rows = []
    for video_id, playlist_index in sorted(playlist_indices.items(), key=lambda item: item[1]):
        if playlist_index < 11:
            continue
        detail = by_id.get(video_id, {})
        rows.append(
            {
                "scope_rank": playlist_index,
                "video_id": video_id,
                "source": "playlist_html",
                "upload_date": detail.get("upload_date", ""),
                "duration": detail.get("duration_string", ""),
                "title": detail.get("title", ""),
                "url": f"https://www.youtube.com/watch?v={video_id}",
                "start_after": "12:20" if video_id == "bDgZhBFm1mU" else "00:00",
                "transcript_path": str(ROOT / "data" / "source" / "craig_youtube" / "transcripts" / f"{video_id}.json"),
                "review_status": "pending",
            }
        )

    live_title = re.compile(r"\bLIVE\b|Live Day Trading|LIVE DAY TRADING|LIVE TRADING CRYPTO", re.I)
    for detail in details:
        if detail.get("upload_date", "") < "20250119":
            continue
        video_id = detail["id"]
        if video_id in playlist_indices:
            continue
        if not live_title.search(detail.get("title", "")):
            continue
        rows.append(
            {
                "scope_rank": "",
                "video_id": video_id,
                "source": "local_details_live_title_not_in_initial_playlist_html",
                "upload_date": detail.get("upload_date", ""),
                "duration": detail.get("duration_string", ""),
                "title": detail.get("title", ""),
                "url": detail.get("url", f"https://www.youtube.com/watch?v={video_id}"),
                "start_after": "00:00",
                "transcript_path": str(ROOT / "data" / "source" / "craig_youtube" / "transcripts" / f"{video_id}.json"),
                "review_status": "pending_playlist_membership_check",
            }
        )

    rows.sort(key=lambda r: (r["upload_date"], str(r["scope_rank"])))
    out_path = OUT_DIR / "review_scope_bdg_forward.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        fieldnames = [
            "scope_rank",
            "video_id",
            "source",
            "upload_date",
            "duration",
            "title",
            "url",
            "start_after",
            "transcript_path",
            "review_status",
        ]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    gold_path = OUT_DIR / "gold_trade_contexts.csv"
    if not gold_path.exists():
        with gold_path.open("w", encoding="utf-8-sig", newline="") as f:
            csv.DictWriter(f, fieldnames=SCHEMA_COLUMNS).writeheader()
    build_manual_seed()
    build_review_scope()


if __name__ == "__main__":
    main()
